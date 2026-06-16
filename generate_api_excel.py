import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "API Endpoints"

# Define headers
headers = ["No", "Modul", "Endpoint (API)", "Method", "Deskripsi"]

# API endpoints data
api_data = [
    # Health Module
    [1, "Health", "/api/health", "GET", "Memeriksa status kesehatan API dan koneksi database"],
    
    # Authentication Module
    [2, "Authentication", "/api/admin/login", "POST", "Melakukan login admin dan mengembalikan CSRF token"],
    [3, "Authentication", "/api/admin/logout", "POST", "Melakukan logout admin dan menghapus session"],
    [4, "Authentication", "/api/csrf-token", "GET", "Mengambil CSRF token untuk validasi request admin"],
    
    # Dashboard Module
    [5, "Dashboard", "/api/dashboard/stats", "GET", "Mengambil statistik dashboard (jumlah data dan bug reports)"],
    
    # Chatbot Module
    [6, "Chatbot", "/api/chat", "POST", "Endpoint chatbot publik untuk pengguna umum"],
    [7, "Chatbot", "/api/admin_chat", "POST", "Endpoint chatbot admin dengan streaming response"],
    
    # Scraped Data Module
    [8, "Scraped Data", "/api/scraped-data", "GET", "Mengambil daftar seluruh data hasil scraping website"],
    [9, "Scraped Data", "/api/scraped-data/{id}", "GET", "Mengambil detail data scraped berdasarkan ID"],
    [10, "Scraped Data", "/api/scraped-data/{id}", "PUT", "Memperbarui data scraped berdasarkan ID"],
    [11, "Scraped Data", "/api/scraped-data/{id}", "DELETE", "Menghapus data scraped berdasarkan ID"],
    
    # Manual Data Module
    [12, "Manual Data", "/api/manual-data", "GET", "Mengambil daftar seluruh data manual yang diupload"],
    [13, "Manual Data", "/api/manual-data/{id}", "GET", "Mengambil detail data manual berdasarkan ID"],
    [14, "Manual Data", "/api/manual-data/{id}", "PUT", "Memperbarui data manual berdasarkan ID"],
    [15, "Manual Data", "/api/manual-data/{id}", "DELETE", "Menghapus data manual berdasarkan ID"],
    [16, "Manual Data", "/api/add_manual_text", "POST", "Menambahkan data manual berupa teks"],
    [17, "Manual Data", "/api/add_manual_file", "POST", "Menambahkan data manual berupa file (PDF, DOCX, PPTX, TXT)"],
    
    # Memory Bank Module
    [18, "Memory Bank", "/api/memory-data", "GET", "Mengambil daftar seluruh data memory bank"],
    [19, "Memory Bank", "/api/memory-data/{id}", "GET", "Mengambil detail data memory berdasarkan ID"],
    [20, "Memory Bank", "/api/memory-data/{id}", "PUT", "Memperbarui data memory berdasarkan ID"],
    [21, "Memory Bank", "/api/memory-data/{id}", "DELETE", "Menghapus data memory berdasarkan ID"],
    [22, "Memory Bank", "/api/save_memory", "POST", "Menyimpan percakapan question-answer ke memory bank"],
    
    # Bug Reports Module
    [23, "Bug Reports", "/api/report_bug", "POST", "Mengirim laporan bug dari pengguna"],
    [24, "Bug Reports", "/api/get_bug_reports", "GET", "Mengambil daftar seluruh laporan bug"],
    [25, "Bug Reports", "/api/bug_reports/{id}", "GET", "Mengambil detail laporan bug berdasarkan ID"],
    [26, "Bug Reports", "/api/bug_reports/{id}/status", "PUT", "Memperbarui status laporan bug"],
    [27, "Bug Reports", "/api/bug_reports/{id}", "DELETE", "Menghapus laporan bug berdasarkan ID"],
    
    # Data Management Module
    [28, "Data Management", "/api/get-data", "GET", "Mengambil semua data (scraped, manual, memory) sekaligus"],
    [29, "Data Management", "/api/data/{type}/{id}", "PUT", "Memperbarui data berdasarkan tipe dan ID"],
    [30, "Data Management", "/api/data/{type}/{id}", "DELETE", "Menghapus data berdasarkan tipe dan ID"],
    
    # System Module
    [31, "System", "/api/scrape", "POST", "Melakukan scraping URL dari file urls_to_scrape.txt"],
    [32, "System", "/api/crawl", "POST", "Melakukan deep crawling website dengan base URL"],
    [33, "System", "/api/reindex", "POST", "Membangun ulang FAISS index untuk vector search"],
    [34, "System", "/api/delete_faiss", "POST", "Menghapus semua direktori FAISS index"],
    [35, "System", "/api/delete_db", "POST", "Mengosongkan semua koleksi database MongoDB"],
]

# Write headers
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Write data
for row_num, row_data in enumerate(api_data, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Apply alternating row colors
        if row_num % 2 == 0:
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

# Set column widths
ws.column_dimensions['A'].width = 5   # No
ws.column_dimensions['B'].width = 18  # Modul
ws.column_dimensions['C'].width = 35  # Endpoint (API)
ws.column_dimensions['D'].width = 10  # Method
ws.column_dimensions['E'].width = 60  # Deskripsi

# Add borders
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws.iter_rows(min_row=1, max_row=len(api_data)+1, min_col=1, max_col=5):
    for cell in row:
        cell.border = thin_border

# Freeze the header row
ws.freeze_panes = "A2"

# Save the workbook
wb.save("API_Endpoints_DamayAI.xlsx")
print("✓ Excel file created successfully: API_Endpoints_DamayAI.xlsx")
print(f"✓ Total endpoints: {len(api_data)}")
